from django.conf import settings

import openpyxl
import requests

from feedback.models import Question, Response

def verify_captcha(token):
    if not token:
        return {"success": False}
    response = requests.post(
        settings.RECAPTCHA_URL,
        data={"secret": settings.RECAPTCHA_KEY,"response": token})
    data = response.json()
    return data

def feedback_to_excel(feedback_qs):
    responses = Response.objects.filter(feedback__in=feedback_qs)
    questions = [
        r.question for r in (
            responses
            .order_by("question__id")
            .distinct("question")
            .select_related("question")
        )
    ]

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Questions"
    question_headers = [
        "Question ID",
        "Text",
        "Type"
    ]
    for i, h in enumerate(question_headers):
        _ = sheet.cell(row=1, column=i+1, value=h)

    for r, q in enumerate(questions):
        _ = sheet.cell(row=r+2, column=1, value=q.id)
        _ = sheet.cell(row=r+2, column=2, value=q.text)
        _ = sheet.cell(row=r+2, column=3, value=q.get_qtype_display())

    response_header = ["Feedback Id", "Timestamp"]
    for q in sorted(questions, key=lambda q: q.id):
        response_header.append(f"Question {q.id} Response")
        if q.qtype == 2:
            response_header.append(f"Question {q.id} Weight")
    sheet = wb.create_sheet("Responses")
    for i, h in enumerate(response_header):
        _ = sheet.cell(row=1, column=i+1, value=h)

    for row, feedback in enumerate(feedback_qs):
        _ = sheet.cell(row=row+2, column=1, value=str(feedback.id))
        _ = sheet.cell(row=row+2, column=2, value=feedback.created.isoformat())
        response_vals = []
        for r in feedback.response_set.order_by('question__id'):
            if r.question.qtype == 1:
                response_vals.append(r.text or "No Response")
            elif r.question.qtype == 2:
                response_vals.append(r.selected.text2 if r.selected else "No Response")
                response_vals.append(r.selected.weight if r.selected else "NaN")
        for col, v in enumerate(response_vals):
            _ = sheet.cell(row=row+2, column=col+3, value=v)

    return wb

        